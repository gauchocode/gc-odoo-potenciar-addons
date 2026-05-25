# -*- coding: utf-8 -*-
import base64
import io
import logging
import re
from uuid import uuid4
from datetime import date, datetime

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

_logger = logging.getLogger(__name__)


class GcSaleInvoiceImportWizard(models.TransientModel):
    _name = "gc.sale.invoice.import.wizard"
    _description = "Importador de facturas de venta desde Excel"
    _preview_max_scan_rows = 20000
    _preview_stop_after_empty_rows = 250
    _analytic_columns = (11, 12, 13, 14)

    file = fields.Binary(string="Archivo Excel", required=True)
    filename = fields.Char(string="Nombre de archivo")
    option_token = fields.Char(
        string="Token opciones hoja",
        default=lambda self: uuid4().hex,
        readonly=True,
    )
    import_type = fields.Selection(
        [
            ("protectores", "Protectores"),
            ("cpd", "CPD"),
            ("prestamos", "Préstamos"),
            ("pagares", "Pagarés"),
            ("on", "ON"),
        ],
        string="Tipo de importación",
        required=True,
        default="protectores",
        help=(
            "Define el comportamiento del importador. "
            "Protectores usa la lógica base del importador; "
            "CPD y Préstamos pueden aplicar comportamientos específicos."
        ),
    )
    sheet_option_ids = fields.Many2many(
        "gc.sale.invoice.import.sheet.option",
        "gc_sale_invoice_import_wizard_sheet_option_rel",
        "wizard_id",
        "option_id",
        string="Hojas disponibles",
        readonly=True,
    )
    sheet_option_id = fields.Many2one(
        "gc.sale.invoice.import.sheet.option",
        string="Hoja a importar",
        required=True,
        help="Seleccione la hoja del Excel a procesar.",
    )
    sheet_name = fields.Char(
        string="Hoja a importar (legacy)",
        required=False,
        help="Nombre exacto de la hoja del Excel a procesar.",
    )
    preview_line_ids = fields.One2many(
        "gc.sale.invoice.import.preview.line",
        "wizard_id",
        string="Preview",
        readonly=True,
    )
    preview_total_count = fields.Integer(
        string="Total",
        compute="_compute_preview_counters",
    )
    preview_ready_count = fields.Integer(
        string="Listas",
        compute="_compute_preview_counters",
    )
    preview_duplicate_count = fields.Integer(
        string="Duplicadas",
        compute="_compute_preview_counters",
    )
    preview_error_count = fields.Integer(
        string="Con error",
        compute="_compute_preview_counters",
    )
    create_partners = fields.Boolean(
        string="Crear clientes faltantes",
        default=True,
        help="Si está activo, se crean clientes que no existan por CUIT.",
    )
    import_log = fields.Text(string="Resultado", readonly=True)

    @api.depends("preview_line_ids.state")
    def _compute_preview_counters(self):
        for wizard in self:
            lines = wizard.preview_line_ids
            wizard.preview_total_count = len(lines)
            wizard.preview_ready_count = len(lines.filtered(lambda line: line.state == "ready"))
            wizard.preview_duplicate_count = len(
                lines.filtered(lambda line: line.state == "duplicate")
            )
            wizard.preview_error_count = len(lines.filtered(lambda line: line.state == "error"))

    @api.onchange("file")
    def _onchange_file_load_sheet_options(self):
        self._clear_sheet_options()
        self.sheet_name = False
        self.sheet_option_id = False
        self.preview_line_ids = [(5, 0, 0)]

        if not self.file or not load_workbook:
            return

        workbook = load_workbook(
            filename=io.BytesIO(base64.b64decode(self.file)),
            read_only=True,
            data_only=True,
        )
        self._set_sheet_options_from_workbook(workbook)

    @api.onchange("sheet_name")
    def _onchange_sheet_name_preview(self):
        self.preview_line_ids = [(5, 0, 0)]
        if not self.file or not self._get_selected_sheet_name() or not load_workbook:
            return
        workbook = load_workbook(
            filename=io.BytesIO(base64.b64decode(self.file)),
            data_only=True,
        )
        self._build_preview_for_selected_sheet(workbook)

    @api.onchange("import_type")
    def _onchange_import_type_preview(self):
        self.preview_line_ids = [(5, 0, 0)]
        if not self.file or not self._get_selected_sheet_name() or not load_workbook:
            return
        workbook = load_workbook(
            filename=io.BytesIO(base64.b64decode(self.file)),
            data_only=True,
        )
        self._build_preview_for_selected_sheet(workbook)

    @api.onchange("sheet_option_id")
    def _onchange_sheet_option_id_sync(self):
        """Mantener sincronía entre selector y nombre de hoja efectivo."""
        if self.sheet_option_id and self.sheet_option_id.name:
            self.sheet_name = self.sheet_option_id.name
            self._onchange_sheet_name_preview()
        else:
            self.sheet_name = False
            self.preview_line_ids = [(5, 0, 0)]

    def action_import_invoices(self):
        self.ensure_one()

        if not load_workbook:
            raise UserError(
                _(
                    "La librería openpyxl no está instalada. "
                    "Instálela en el contenedor/entorno de Odoo."
                )
            )
        if not self.file:
            raise UserError(_("Debe seleccionar un archivo."))

        workbook = load_workbook(
            filename=io.BytesIO(base64.b64decode(self.file)),
            data_only=True,
        )
        selected_sheet_name = self._get_selected_sheet_name()
        if not selected_sheet_name:
            raise UserError(_("Debe seleccionar la hoja a importar."))
        if selected_sheet_name not in workbook.sheetnames:
            raise UserError(
                _("La hoja seleccionada '%s' no existe en el archivo.")
                % selected_sheet_name
            )
        sheet = workbook[selected_sheet_name]
        if not self.preview_line_ids:
            self._build_preview_for_selected_sheet(workbook)

        row_numbers = sorted(
            self.preview_line_ids.filtered(lambda line: line.state == "ready").mapped("row_number")
        )
        if not row_numbers:
            raise UserError(_("No hay filas válidas para importar en la hoja seleccionada."))

        headers = {
            column: (sheet.cell(row=1, column=column).value or "").strip()
            for column in range(1, 22)
        }
        product_code_by_column = {
            17: self._extract_code_from_header(headers.get(17)),
            18: self._extract_code_from_header(headers.get(18)),
            19: self._extract_code_from_header(headers.get(19)),
        }

        created, skipped, errors = 0, 0, 0
        log_lines = []

        for row_idx in row_numbers:
            try:
                evaluation = self._evaluate_row_by_type(sheet, row_idx)
                row = evaluation.get("row")
                state = evaluation.get("state")

                if state == "skip":
                    skipped += 1
                    continue
                if state == "duplicate":
                    skipped += 1
                    log_lines.append(
                        _(
                            "⊘ Fila %(row)s: omitida por duplicado (Factura ID %(move)s)."
                        )
                        % {"row": row_idx, "move": evaluation["duplicate_move_id"]}
                    )
                    continue
                if state == "error":
                    errors += 1
                    log_lines.append(
                        _("✗ Fila %(row)s: %(error)s")
                        % {
                            "row": row_idx,
                            "error": evaluation.get("message") or _("Error inesperado"),
                        }
                    )
                    continue

                move = self._create_invoice_by_type(row, headers, product_code_by_column)
                created += 1
                log_lines.append(
                    _(
                        "✓ Fila %(row)s: factura borrador %(move)s creada para %(partner)s por %(total).2f."
                    )
                    % {
                        "row": row_idx,
                        "move": move.id,
                        "partner": move.partner_id.display_name,
                        "total": row["total"],
                    }
                )
            except Exception as error:
                errors += 1
                message = _("✗ Fila %(row)s: %(error)s") % {
                    "row": row_idx,
                    "error": str(error),
                }
                log_lines.append(message)
                _logger.exception("Error importando fila %s", row_idx)

        self.import_log = "\n".join(
            [
                "====================",
                "RESUMEN IMPORTACIÓN",
                "====================",
                _("Tipo de importación: %s") % dict(self._fields["import_type"].selection).get(
                    self.import_type, self.import_type
                ),
                _("Hoja procesada: %s") % sheet.title,
                _("Filas analizadas: %s") % max(sheet.max_row - 1, 0),
                _("Facturas creadas: %s") % created,
                _("Filas omitidas: %s") % skipped,
                _("Errores: %s") % errors,
                "",
            ]
            + log_lines
        )

        return {
            "type": "ir.actions.act_window",
            "res_model": "gc.sale.invoice.import.wizard",
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
            "context": {"show_log": True},
        }

    def _clear_sheet_options(self):
        self.ensure_one()
        if self.option_token:
            options = self.env["gc.sale.invoice.import.sheet.option"].sudo().search(
                [("token", "=", self.option_token)]
            )
            options.unlink()
        self.sheet_option_ids = [(5, 0, 0)]

    def _set_sheet_options_from_workbook(self, workbook):
        self.ensure_one()
        token = self.option_token or uuid4().hex
        self.option_token = token

        Option = self.env["gc.sale.invoice.import.sheet.option"].sudo()
        old_options = Option.search([("token", "=", token)])
        old_options.unlink()

        names = workbook.sheetnames
        if not names:
            self.sheet_option_ids = [(5, 0, 0)]
            self.sheet_option_id = False
            self.sheet_name = False
            return False

        created = Option.create([{"name": name, "token": token} for name in names])
        self.sheet_option_ids = [(6, 0, created.ids)]
        self.sheet_option_id = False
        self.sheet_name = False
        return bool(created)

    def _build_preview_for_selected_sheet(self, workbook):
        self.ensure_one()
        selected_sheet_name = self._get_selected_sheet_name()
        if not selected_sheet_name or selected_sheet_name not in workbook.sheetnames:
            self.preview_line_ids = [(5, 0, 0)]
            return

        sheet = workbook[selected_sheet_name]
        commands = [(5, 0, 0)]

        for row_idx in self._iter_preview_row_numbers(sheet):
            try:
                evaluation = self._evaluate_row_by_type(sheet, row_idx)
                state = evaluation.get("state")
                parsed = evaluation.get("row")

                if state == "skip" or not parsed:
                    continue

                commands.append(
                    (
                        0,
                        0,
                        {
                            "row_number": row_idx,
                            "state": state,
                            "message": evaluation.get("message"),
                            "invoice_date": parsed.get("invoice_date"),
                            "cuit": parsed.get("cuit"),
                            "partner_name": parsed.get("partner_name"),
                            "comment": parsed.get("comment"),
                            "journal_name": parsed.get("journal_name"),
                            "total": parsed.get("total"),
                            "line_amount_17": parsed.get("line_amount_17"),
                            "line_amount_18": parsed.get("line_amount_18"),
                            "line_amount_19": parsed.get("line_amount_19"),
                            "vat_21": parsed.get("vat_21"),
                        },
                    )
                )
            except Exception as error:
                commands.append(
                    (
                        0,
                        0,
                        self._build_preview_error_vals(sheet, row_idx, str(error)),
                    )
                )

        self.preview_line_ids = commands

    def _get_selected_sheet_name(self):
        return (self.sheet_option_id.name if self.sheet_option_id else False) or self.sheet_name

    def _parse_row_by_type(self, sheet, row_idx):
        parser = self._get_type_method("parse_row")
        if not parser:
            raise ValidationError(
                _("No hay parser configurado para el tipo de importación '%s'.")
                % self.import_type
            )
        return parser(sheet, row_idx)

    def _create_invoice_by_type(self, row, headers, product_code_by_column):
        creator = self._get_type_method("create_invoice")
        if not creator:
            raise ValidationError(
                _("No hay creador configurado para el tipo de importación '%s'.")
                % self.import_type
            )
        return creator(row, headers, product_code_by_column)

    def _find_duplicated_invoice_by_type(self, row):
        finder = self._get_type_method("find_duplicate")
        if not finder:
            raise ValidationError(
                _("No hay validador de duplicados para el tipo de importación '%s'.")
                % self.import_type
            )
        return finder(row)

    def _get_type_method(self, base_name):
        return getattr(self, "_%s_type_%s" % (base_name, self.import_type), None)

    def _evaluate_row_by_type(self, sheet, row_idx):
        try:
            parsed = self._parse_row_by_type(sheet, row_idx)
            if not parsed:
                return {"state": "skip", "row": None, "message": False}

            for analytic_item in parsed.get("analytic_labels", []):
                self._find_analytic_account(
                    analytic_item["label"],
                    plan_label=self._cell(sheet, 1, analytic_item["column"]),
                    column=analytic_item["column"],
                )

            self._validate_row_references_by_type(sheet, parsed)

            duplicated = self._find_duplicated_invoice_by_type(parsed)
            if duplicated:
                return {
                    "state": "duplicate",
                    "row": parsed,
                    "message": _("Factura existente ID %s") % duplicated.id,
                    "duplicate_move_id": duplicated.id,
                }

            return {"state": "ready", "row": parsed, "message": False}
        except Exception as error:
            return {
                "state": "error",
                "row": self._fallback_row_data_for_preview(sheet, row_idx),
                "message": str(error),
            }

    def _validate_row_references_by_type(self, sheet, row):
        validator = self._get_type_method("validate_references")
        if validator:
            return validator(sheet, row)
        return True

    # --- Hooks por tipo ---
    def _parse_row_type_protectores(self, sheet, row_idx):
        return self._parse_row(sheet, row_idx)

    def _parse_row_type_cpd(self, sheet, row_idx):
        first_cell = self._cell(sheet, row_idx, 1)
        if isinstance(first_cell, str) and first_cell.strip().lower() == "fecha fc":
            return None

        row = self._parse_row(sheet, row_idx)
        if not row:
            return row

        row["import_type"] = "cpd"
        self._validate_cpd_total(row)
        return row

    def _parse_row_type_prestamos(self, sheet, row_idx):
        first_cell = self._cell(sheet, row_idx, 1)
        if isinstance(first_cell, str) and first_cell.strip().lower() == "fecha fc":
            return None

        values = [self._cell(sheet, row_idx, column) for column in range(1, 22)]
        fecha_fc = values[0]
        cuit = values[3]
        partner_name = values[4]
        total = self._to_float(values[20])  # col 21

        if not any(values[:21]):
            return None
        if not (fecha_fc and cuit and partner_name and total):
            return None

        invoice_date = self._to_date(fecha_fc)
        cuit_clean = re.sub(r"\D", "", str(cuit or ""))
        if len(cuit_clean) < 8:
            raise ValidationError(_("CUIT inválido: %s") % cuit)

        return {
            "invoice_date": invoice_date,
            "period_from": self._to_date(values[1], allow_empty=True),
            "period_to": self._to_date(values[2], allow_empty=True),
            "cuit": cuit_clean,
            "partner_name": str(partner_name).strip(),
            "comment": str(values[5]).strip() if values[5] else False,
            "exchange_rate": self._to_float(values[6], allow_empty=True),
            "currency_label": str(values[7]).strip() if values[7] else False,
            "journal_name": str(values[8]).strip() if values[8] else False,
            "payment_term_name": str(values[9]).strip() if values[9] else False,
            "analytic_label": str(values[10]).strip() if values[10] else False,
            "analytic_labels": self._get_analytic_labels_from_values(values),
            "base_total": self._to_float(values[14], allow_empty=True) or 0.0,
            "pct": self._to_float(values[15], allow_empty=True),
            "line_amount_17": self._to_float(values[16], allow_empty=True) or 0.0,
            "line_amount_18": self._to_float(values[17], allow_empty=True) or 0.0,
            "line_amount_19": self._to_float(values[18], allow_empty=True) or 0.0,
            "vat_21": self._to_float(values[19], allow_empty=True) or 0.0,
            "total": total,
            "import_type": "prestamos",
        }

    def _create_invoice_type_protectores(self, row, headers, product_code_by_column):
        return self._create_invoice(row, headers, product_code_by_column)

    def _create_invoice_type_cpd(self, row, headers, product_code_by_column):
        self._validate_cpd_total(row)

        move_model = self.env["account.move"]
        partner = self._get_or_create_partner(row)
        journal = self._get_sale_journal(row.get("journal_name"))
        currency = self._get_currency(row.get("currency_label"))
        tax_21 = self._get_sale_tax_21()
        if not tax_21:
            raise ValidationError(_("No se encontró un impuesto de venta del 21% para CPD."))
        tax_exempt = self._get_sale_tax_exempt()
        if not tax_exempt:
            raise ValidationError(_("No se encontró el impuesto de venta IVA EXENTO para CPD."))
        analytic_accounts = self._find_analytic_accounts(row, headers)

        mapped_codes = dict(product_code_by_column or {})
        mapped_codes[17] = mapped_codes.get(17) or "000005"
        mapped_codes[18] = mapped_codes.get(18) or "000012"

        lines = []
        if row.get("comment"):
            lines.append((0, 0, self._build_comment_line_vals(row["comment"])))

        exempt_amount = row.get("line_amount_17") or 0.0
        taxed_amount = row.get("line_amount_18") or 0.0

        if exempt_amount > 0:
            exempt_line_vals = self._build_line_vals(
                amount=exempt_amount,
                column=17,
                header=headers.get(17),
                product_code=mapped_codes.get(17),
                tax_21=False,
                analytic_accounts=analytic_accounts,
            )
            # CPD: la comisión de columna Q debe llevar explícitamente IVA EXENTO,
            # aunque el producto tenga otros impuestos de venta por defecto.
            exempt_line_vals["tax_ids"] = [(6, 0, [tax_exempt.id])]
            lines.append(
                (
                    0,
                    0,
                    exempt_line_vals,
                )
            )
        if taxed_amount > 0:
            lines.append(
                (
                    0,
                    0,
                    self._build_line_vals(
                        amount=taxed_amount,
                        column=18,
                        header=headers.get(18),
                        product_code=mapped_codes.get(18),
                        tax_21=tax_21,
                        analytic_accounts=analytic_accounts,
                    ),
                )
            )
        if exempt_amount <= 0 and taxed_amount <= 0:
            raise ValidationError(
                _("No se pudieron determinar importes CPD para las columnas Q/R.")
            )

        origin = False
        if row.get("period_from") and row.get("period_to"):
            origin = _(
                "Período: %(from)s a %(to)s"
            ) % {
                "from": row["period_from"].strftime("%Y-%m-%d"),
                "to": row["period_to"].strftime("%Y-%m-%d"),
            }

        vals = {
            "move_type": "out_invoice",
            "partner_id": partner.id,
            "invoice_date": row["invoice_date"],
            "date": row["invoice_date"],
            "journal_id": journal.id,
            "currency_id": currency.id,
            "ref": row.get("comment") or False,
            "invoice_origin": origin,
            # Mantener comportamiento existente: no setear condición de cobro
            # desde Excel; se hereda/gestiona en el contacto.
            "invoice_line_ids": lines,
            "narration": row.get("comment") or False,
        }

        if row.get("period_from") and "l10n_ar_afip_service_start" in move_model._fields:
            vals["l10n_ar_afip_service_start"] = row["period_from"]
        if row.get("period_to") and "l10n_ar_afip_service_end" in move_model._fields:
            vals["l10n_ar_afip_service_end"] = row["period_to"]
        if row.get("exchange_rate") and "l10n_ar_currency_rate" in move_model._fields:
            vals["l10n_ar_currency_rate"] = row["exchange_rate"]

        return move_model.create(vals)

    def _create_invoice_type_prestamos(self, row, headers, product_code_by_column):
        move_model = self.env["account.move"]
        partner = self._get_or_create_partner(row)
        journal = self._get_sale_journal(row.get("journal_name"))
        currency = self._get_currency(row.get("currency_label"))
        tax_21 = self._get_sale_tax_21()
        if not tax_21:
            raise ValidationError(_("No se encontró un impuesto de venta del 21% para Préstamos."))
        analytic_accounts = self._find_analytic_accounts(row, headers)

        mapped_codes = dict(product_code_by_column or {})
        mapped_codes[17] = mapped_codes.get(17) or "000003"
        mapped_codes[18] = mapped_codes.get(18) or "000004"

        lines = []
        if row.get("comment"):
            lines.append((0, 0, self._build_comment_line_vals(row["comment"])))

        amount_lines = 0
        for column in (17, 18, 19):
            amount = row.get(f"line_amount_{column}") or 0.0
            if amount <= 0:
                continue
            lines.append(
                (
                    0,
                    0,
                    self._build_line_vals(
                        amount=amount,
                        column=column,
                        header=headers.get(column),
                        product_code=mapped_codes.get(column),
                        tax_21=tax_21,
                        analytic_accounts=analytic_accounts,
                    ),
                )
            )
            amount_lines += 1

        if not amount_lines:
            raise ValidationError(
                _("No se pudieron determinar importes de Préstamos para las columnas Q/R/S.")
            )

        origin = False
        if row.get("period_from") and row.get("period_to"):
            origin = _(
                "Período: %(from)s a %(to)s"
            ) % {
                "from": row["period_from"].strftime("%Y-%m-%d"),
                "to": row["period_to"].strftime("%Y-%m-%d"),
            }

        vals = {
            "move_type": "out_invoice",
            "partner_id": partner.id,
            "invoice_date": row["invoice_date"],
            "date": row["invoice_date"],
            "journal_id": journal.id,
            "currency_id": currency.id,
            "ref": row.get("comment") or False,
            "invoice_origin": origin,
            "invoice_line_ids": lines,
            "narration": row.get("comment") or False,
        }

        if row.get("period_from") and "l10n_ar_afip_service_start" in move_model._fields:
            vals["l10n_ar_afip_service_start"] = row["period_from"]
        if row.get("period_to") and "l10n_ar_afip_service_end" in move_model._fields:
            vals["l10n_ar_afip_service_end"] = row["period_to"]
        if row.get("exchange_rate") and "l10n_ar_currency_rate" in move_model._fields:
            vals["l10n_ar_currency_rate"] = row["exchange_rate"]

        return move_model.create(vals)

    def _find_duplicate_type_protectores(self, row):
        return self._find_duplicated_invoice(row)

    def _find_duplicate_type_cpd(self, row):
        return self._find_duplicated_invoice(row)

    def _find_duplicate_type_prestamos(self, row):
        return self._find_duplicated_invoice(row)

    def _parse_row_type_pagares(self, sheet, row_idx):
        return self._parse_row_type_prestamos(sheet, row_idx)

    def _create_invoice_type_pagares(self, row, headers, product_code_by_column):
        move_model = self.env["account.move"]
        partner = self._get_or_create_partner(row)
        journal = self._get_sale_journal(row.get("journal_name"))
        currency = self._get_currency(row.get("currency_label"))
        tax_21 = self._get_sale_tax_21()
        if not tax_21:
            raise ValidationError(_("No se encontró un impuesto de venta del 21% para Pagarés."))
        tax_exempt = self._get_sale_tax_exempt()
        if not tax_exempt:
            raise ValidationError(_("No se encontró el impuesto de venta IVA EXENTO para Pagarés."))
        analytic_accounts = self._find_analytic_accounts(row, headers)

        lines = []
        if row.get("comment"):
            lines.append((0, 0, self._build_comment_line_vals(row["comment"])))

        amount_lines = 0

        # Col 17: exento (IVA EXENTO)
        exempt_amount = row.get("line_amount_17") or 0.0
        if exempt_amount > 0:
            exempt_vals = self._build_line_vals(
                amount=exempt_amount,
                column=17,
                header=headers.get(17),
                product_code=product_code_by_column.get(17),
                tax_21=False,
                analytic_accounts=analytic_accounts,
            )
            exempt_vals["tax_ids"] = [(6, 0, [tax_exempt.id])]
            lines.append((0, 0, exempt_vals))
            amount_lines += 1

        # Cols 18 y 19: gravados (IVA 21%)
        for column in (18, 19):
            amount = row.get(f"line_amount_{column}") or 0.0
            if amount <= 0:
                continue
            lines.append(
                (
                    0,
                    0,
                    self._build_line_vals(
                        amount=amount,
                        column=column,
                        header=headers.get(column),
                        product_code=product_code_by_column.get(column),
                        tax_21=tax_21,
                        analytic_accounts=analytic_accounts,
                    ),
                )
            )
            amount_lines += 1

        if not amount_lines:
            raise ValidationError(
                _("No se pudieron determinar importes de Pagarés para las columnas Q/R/S.")
            )

        origin = False
        if row.get("period_from") and row.get("period_to"):
            origin = _(
                "Período: %(from)s a %(to)s"
            ) % {
                "from": row["period_from"].strftime("%Y-%m-%d"),
                "to": row["period_to"].strftime("%Y-%m-%d"),
            }

        vals = {
            "move_type": "out_invoice",
            "partner_id": partner.id,
            "invoice_date": row["invoice_date"],
            "date": row["invoice_date"],
            "journal_id": journal.id,
            "currency_id": currency.id,
            "ref": row.get("comment") or False,
            "invoice_origin": origin,
            "invoice_line_ids": lines,
            "narration": row.get("comment") or False,
        }

        if row.get("period_from") and "l10n_ar_afip_service_start" in move_model._fields:
            vals["l10n_ar_afip_service_start"] = row["period_from"]
        if row.get("period_to") and "l10n_ar_afip_service_end" in move_model._fields:
            vals["l10n_ar_afip_service_end"] = row["period_to"]
        if row.get("exchange_rate") and "l10n_ar_currency_rate" in move_model._fields:
            vals["l10n_ar_currency_rate"] = row["exchange_rate"]

        return move_model.create(vals)

    def _find_duplicate_type_pagares(self, row):
        return self._find_duplicated_invoice(row)

    def _parse_row_type_on(self, sheet, row_idx):
        first_cell = self._cell(sheet, row_idx, 1)
        if isinstance(first_cell, str) and first_cell.strip().lower() == "fecha fc":
            return None
        return self._parse_row(sheet, row_idx)

    def _create_invoice_type_on(self, row, headers, product_code_by_column):
        return self._create_invoice_type_cpd(row, headers, product_code_by_column)

    def _find_duplicate_type_on(self, row):
        return self._find_duplicated_invoice(row)

    def _validate_references_type_on(self, sheet, row):
        return self._validate_references_type_cpd(sheet, row)

    def _validate_references_type_pagares(self, sheet, row):
        headers = {
            column: self._cell(sheet, 1, column)
            for column in range(1, 22)
        }
        product_code_by_column = {
            17: self._extract_code_from_header(headers.get(17)),
            18: self._extract_code_from_header(headers.get(18)),
            19: self._extract_code_from_header(headers.get(19)),
        }
        self._get_sale_journal(row.get("journal_name"))
        self._get_currency(row.get("currency_label"))
        if not self._get_sale_tax_21():
            raise ValidationError(_("No se encontró un impuesto de venta del 21% para Pagarés."))
        if not self._get_sale_tax_exempt():
            raise ValidationError(_("No se encontró el impuesto de venta IVA EXENTO para Pagarés."))
        for column, code in product_code_by_column.items():
            if code and not self._find_product(code):
                raise ValidationError(
                    _("No se encontró el producto Pagarés [%(code)s] (columna %(column)s).")
                    % {"code": code, "column": self._column_number_to_letter(column)}
                )
        return True

    def _validate_references_type_cpd(self, sheet, row):
        headers = {
            column: self._cell(sheet, 1, column)
            for column in range(1, 21)
        }
        product_code_by_column = {
            17: self._extract_code_from_header(headers.get(17)) or "000005",
            18: self._extract_code_from_header(headers.get(18)) or "000012",
        }
        self._get_sale_journal(row.get("journal_name"))
        self._get_currency(row.get("currency_label"))
        if not self._get_sale_tax_21():
            raise ValidationError(_("No se encontró un impuesto de venta del 21% para CPD."))
        if not self._get_sale_tax_exempt():
            raise ValidationError(_("No se encontró el impuesto de venta IVA EXENTO para CPD."))
        for column, code in product_code_by_column.items():
            if not self._find_product(code):
                raise ValidationError(
                    _("No se encontró el producto CPD [%(code)s] (columna %(column)s).")
                    % {"code": code, "column": self._column_number_to_letter(column)}
                )
        self._validate_cpd_total(row)
        return True

    def _validate_references_type_prestamos(self, sheet, row):
        headers = {
            column: self._cell(sheet, 1, column)
            for column in range(1, 22)
        }
        product_code_by_column = {
            17: self._extract_code_from_header(headers.get(17)),
            18: self._extract_code_from_header(headers.get(18)),
            19: self._extract_code_from_header(headers.get(19)),
        }
        self._get_sale_journal(row.get("journal_name"))
        self._get_currency(row.get("currency_label"))
        if not self._get_sale_tax_21():
            raise ValidationError(_("No se encontró un impuesto de venta del 21% para Préstamos."))
        mapped_codes = dict(product_code_by_column or {})
        mapped_codes[17] = mapped_codes.get(17) or "000003"
        mapped_codes[18] = mapped_codes.get(18) or "000004"
        for column, code in mapped_codes.items():
            if code and not self._find_product(code):
                raise ValidationError(
                    _("No se encontró el producto Préstamos [%(code)s] (columna %(column)s).")
                    % {"code": code, "column": self._column_number_to_letter(column)}
                )
        return True

    def _validate_cpd_total(self, row):
        expected_total = (row.get("line_amount_17") or 0.0) + (
            row.get("line_amount_18") or 0.0
        ) + (row.get("vat_21") or 0.0)
        total = row.get("total") or 0.0
        currency = self._get_currency(row.get("currency_label"))
        tolerance = max(currency.rounding or 0.01, 0.01)
        if abs(expected_total - total) > tolerance:
            raise ValidationError(
                _(
                    "El total CPD no coincide: exento + gravado + IVA = "
                    "%(expected).2f, total informado = %(total).2f."
                )
                % {"expected": expected_total, "total": total}
            )
        return True

    def _iter_preview_row_numbers(self, sheet):
        max_row = min(sheet.max_row or 0, self._preview_max_scan_rows)
        empty_streak = 0
        for row_idx in range(2, max_row + 1):
            if self._row_has_key_data(sheet, row_idx):
                empty_streak = 0
                yield row_idx
                continue

            empty_streak += 1
            if empty_streak >= self._preview_stop_after_empty_rows:
                break

    def _row_has_key_data(self, sheet, row_idx):
        # Columnas mínimas para considerar fila candidata.
        key_columns = (1, 4, 5, 20)
        return any(self._cell(sheet, row_idx, col) not in (None, "") for col in key_columns)

    def _fallback_row_data_for_preview(self, sheet, row_idx):
        if self.import_type in ("prestamos", "pagares"):
            return {
                "invoice_date": False,
                "cuit": str(self._cell(sheet, row_idx, 4) or ""),
                "partner_name": str(self._cell(sheet, row_idx, 5) or ""),
                "comment": str(self._cell(sheet, row_idx, 6) or ""),
                "journal_name": str(self._cell(sheet, row_idx, 9) or ""),
                "total": self._to_float(self._cell(sheet, row_idx, 21), allow_empty=True) or 0.0,
                "line_amount_17": self._to_float(self._cell(sheet, row_idx, 17), allow_empty=True) or 0.0,
                "line_amount_18": self._to_float(self._cell(sheet, row_idx, 18), allow_empty=True) or 0.0,
                "line_amount_19": self._to_float(self._cell(sheet, row_idx, 19), allow_empty=True) or 0.0,
                "vat_21": self._to_float(self._cell(sheet, row_idx, 20), allow_empty=True) or 0.0,
            }
        return {
            "invoice_date": False,
            "cuit": str(self._cell(sheet, row_idx, 4) or ""),
            "partner_name": str(self._cell(sheet, row_idx, 5) or ""),
            "comment": str(self._cell(sheet, row_idx, 6) or ""),
            "journal_name": str(self._cell(sheet, row_idx, 9) or ""),
            "total": self._to_float(self._cell(sheet, row_idx, 20), allow_empty=True) or 0.0,
            "line_amount_17": self._to_float(self._cell(sheet, row_idx, 17), allow_empty=True) or 0.0,
            "line_amount_18": self._to_float(self._cell(sheet, row_idx, 18), allow_empty=True) or 0.0,
            "vat_21": self._to_float(self._cell(sheet, row_idx, 19), allow_empty=True) or 0.0,
        }

    def _build_preview_error_vals(self, sheet, row_idx, message):
        vals = self._fallback_row_data_for_preview(sheet, row_idx)
        vals.update(
            {
                "row_number": row_idx,
                "state": "error",
                "message": message,
            }
        )
        return vals

    def _parse_row(self, sheet, row_idx):
        values = [self._cell(sheet, row_idx, column) for column in range(1, 21)]

        fecha_fc = values[0]
        cuit = values[3]
        partner_name = values[4]
        total = self._to_float(values[19])

        # Ignorar filas vacías/de relleno.
        if not any(values[:20]):
            return None
        if not (fecha_fc and cuit and partner_name and total):
            return None

        invoice_date = self._to_date(fecha_fc)
        cuit_clean = re.sub(r"\D", "", str(cuit or ""))
        if len(cuit_clean) < 8:
            raise ValidationError(_("CUIT inválido: %s") % cuit)

        return {
            "invoice_date": invoice_date,
            "period_from": self._to_date(values[1], allow_empty=True),
            "period_to": self._to_date(values[2], allow_empty=True),
            "cuit": cuit_clean,
            "partner_name": str(partner_name).strip(),
            "comment": str(values[5]).strip() if values[5] else False,
            "exchange_rate": self._to_float(values[6], allow_empty=True),
            "currency_label": str(values[7]).strip() if values[7] else False,
            "journal_name": str(values[8]).strip() if values[8] else False,
            "payment_term_name": str(values[9]).strip() if values[9] else False,
            "analytic_label": str(values[10]).strip() if values[10] else False,
            "analytic_labels": self._get_analytic_labels_from_values(values),
            "base_total": self._to_float(values[14], allow_empty=True) or 0.0,
            "pct": self._to_float(values[15], allow_empty=True),
            "line_amount_17": self._to_float(values[16], allow_empty=True) or 0.0,
            "line_amount_18": self._to_float(values[17], allow_empty=True) or 0.0,
            "vat_21": self._to_float(values[18], allow_empty=True) or 0.0,
            "total": total,
        }

    def _create_invoice(self, row, headers, product_code_by_column):
        move_model = self.env["account.move"]
        partner = self._get_or_create_partner(row)
        journal = self._get_sale_journal(row.get("journal_name"))
        currency = self._get_currency(row.get("currency_label"))
        tax_21 = self._get_sale_tax_21()
        analytic_accounts = self._find_analytic_accounts(row, headers)

        lines = []
        if row.get("comment"):
            lines.append((0, 0, self._build_comment_line_vals(row["comment"])))

        amount_lines = 0
        for column in (17, 18):
            amount = row.get(f"line_amount_{column}") or 0.0
            if amount <= 0:
                continue
            line_vals = self._build_line_vals(
                amount=amount,
                column=column,
                header=headers.get(column),
                product_code=product_code_by_column.get(column),
                tax_21=tax_21,
                analytic_accounts=analytic_accounts,
            )
            lines.append((0, 0, line_vals))
            amount_lines += 1

        if not amount_lines:
            fallback_base = row["base_total"] or (row["total"] - row["vat_21"])
            if fallback_base <= 0:
                raise ValidationError(
                    _("No se pudieron determinar importes de líneas para la factura.")
                )
            lines.append(
                (
                    0,
                    0,
                    self._build_line_vals(
                        amount=fallback_base,
                        column=15,
                        header=_("Comisión Total"),
                        product_code=False,
                        tax_21=tax_21,
                        analytic_accounts=analytic_accounts,
                    ),
                )
            )

        origin = False
        if row.get("period_from") and row.get("period_to"):
            origin = _(
                "Período: %(from)s a %(to)s"
            ) % {
                "from": row["period_from"].strftime("%Y-%m-%d"),
                "to": row["period_to"].strftime("%Y-%m-%d"),
            }

        vals = {
            "move_type": "out_invoice",
            "partner_id": partner.id,
            "invoice_date": row["invoice_date"],
            "date": row["invoice_date"],
            "journal_id": journal.id,
            "currency_id": currency.id,
            "ref": row.get("comment") or False,
            "invoice_origin": origin,
            # Importante: no setear condición de cobro desde Excel en la factura.
            # En este proyecto la condición se gestiona en el contacto
            # (campo payment_condition_id) y se hereda al asignar partner.
            "invoice_line_ids": lines,
            "narration": row.get("comment") or False,
        }

        # B/C del Excel: período de facturación AFIP (desde/hasta).
        # Se asigna solo si los campos existen en account.move (compatibilidad).
        if row.get("period_from") and "l10n_ar_afip_service_start" in move_model._fields:
            vals["l10n_ar_afip_service_start"] = row["period_from"]
        if row.get("period_to") and "l10n_ar_afip_service_end" in move_model._fields:
            vals["l10n_ar_afip_service_end"] = row["period_to"]

        if row.get("exchange_rate") and "l10n_ar_currency_rate" in move_model._fields:
            vals["l10n_ar_currency_rate"] = row["exchange_rate"]

        return move_model.create(vals)

    def _build_comment_line_vals(self, comment):
        return {
            "display_type": "line_note",
            "name": comment,
        }

    def _build_line_vals(
        self, amount, column, header, product_code, tax_21, analytic_accounts=False
    ):
        product = self._find_product(product_code)
        line_name = header or _("Línea %s") % column

        vals = {
            "name": line_name,
            "quantity": 1.0,
            "price_unit": amount,
        }
        if product:
            vals["product_id"] = product.id
            vals["name"] = product.display_name
        else:
            account = self._get_default_income_account()
            if not account:
                raise ValidationError(
                    _(
                        "No se encontró una cuenta de ingresos para la línea '%s'. "
                        "Configure una cuenta de ingresos."
                    )
                    % line_name
                )
            vals["account_id"] = account.id
        if tax_21:
            vals["tax_ids"] = [(6, 0, [tax_21.id])]
        if analytic_accounts:
            line_fields = self.env["account.move.line"]._fields
            if "analytic_distribution" in line_fields:
                vals["analytic_distribution"] = {
                    str(account.id): 100.0 for account in analytic_accounts
                }
            elif "analytic_account_id" in line_fields:
                vals["analytic_account_id"] = analytic_accounts[:1].id
        return vals

    def _get_analytic_labels_from_values(self, values):
        labels = []
        for column in self._analytic_columns:
            value = values[column - 1]
            if value:
                labels.append(
                    {
                        "column": column,
                        "label": str(value).strip(),
                    }
                )
        return labels

    def _find_analytic_accounts(self, row, headers):
        AnalyticAccount = self.env["account.analytic.account"]
        accounts = AnalyticAccount.browse()
        for analytic_item in row.get("analytic_labels", []):
            account = self._find_analytic_account(
                analytic_item["label"],
                plan_label=headers.get(analytic_item["column"]),
                column=analytic_item["column"],
            )
            if account and account.id not in accounts.ids:
                accounts |= account
        return accounts

    def _find_analytic_account(self, analytic_label, plan_label=False, column=False):
        if not analytic_label:
            return False

        AnalyticAccount = self.env["account.analytic.account"]
        label = str(analytic_label).strip()
        plan = self._find_analytic_plan(plan_label)
        plan_domain = []
        if plan and "plan_id" in AnalyticAccount._fields:
            plan_domain = [("plan_id", "=", plan.id)]

        bracket_code = self._extract_code_from_header(label)
        candidates = [label]
        if bracket_code:
            candidates.append(bracket_code)

        if label.startswith("[") and "]" in label:
            name_part = label.split("]", 1)[1].strip()
            if name_part:
                candidates.append(name_part)

        exact_domains = []
        for candidate in candidates:
            exact_domains.append(plan_domain + [("name", "=", candidate)])
            if "code" in AnalyticAccount._fields:
                exact_domains.append(plan_domain + [("code", "=", candidate)])

        for domain in exact_domains:
            account = AnalyticAccount.search(domain, limit=1)
            if account:
                return account

        for candidate in candidates:
            matches = AnalyticAccount.name_search(
                name=candidate,
                args=plan_domain,
                operator="ilike",
                limit=2,
            )
            if len(matches) == 1:
                return AnalyticAccount.browse(matches[0][0])

        column_label = self._column_number_to_letter(column) if column else _("K-N")
        if plan:
            raise ValidationError(
                _(
                    "No se encontró la cuenta analítica/etiqueta '%(label)s' "
                    "dentro de '%(plan)s' (columna %(column)s)."
                )
                % {
                    "label": analytic_label,
                    "plan": plan.display_name,
                    "column": column_label,
                }
            )
        raise ValidationError(
            _(
                "No se encontró la cuenta analítica/etiqueta de la columna "
                "%(column)s: %(label)s"
            )
            % {"column": column_label, "label": analytic_label}
        )

    def _find_analytic_plan(self, plan_label):
        if not plan_label:
            return False

        try:
            AnalyticPlan = self.env["account.analytic.plan"]
        except KeyError:
            return False

        candidates = self._get_analytic_plan_candidates(plan_label)
        for candidate in candidates:
            plan = AnalyticPlan.search([("name", "=", candidate)], limit=1)
            if plan:
                return plan
            if "code" in AnalyticPlan._fields:
                plan = AnalyticPlan.search([("code", "=", candidate)], limit=1)
                if plan:
                    return plan

        for candidate in candidates:
            plan = AnalyticPlan.search([("name", "ilike", candidate)], limit=2)
            if len(plan) == 1:
                return plan
        return False

    @staticmethod
    def _get_analytic_plan_candidates(plan_label):
        label = str(plan_label or "").strip()
        if not label:
            return []

        short_label = re.sub(r"(?i)^anal[ií]tico\s*", "", label).strip()
        candidates = [label]
        if short_label and short_label not in candidates:
            candidates.append(short_label)
        return candidates

    @staticmethod
    def _column_number_to_letter(column):
        if not column:
            return ""
        letters = ""
        while column:
            column, remainder = divmod(column - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    def _find_product(self, product_code):
        if product_code:
            product = self.env["product.product"].search(
                [("default_code", "=", product_code)],
                limit=1,
            )
            if product:
                return product
        return False

    def _find_duplicated_invoice(self, row):
        if not row.get("comment"):
            return False
        partner = self.env["res.partner"].search(
            [("vat", "=", row["cuit"])],
            limit=1,
        )
        if not partner:
            return False
        return self.env["account.move"].search(
            [
                ("move_type", "=", "out_invoice"),
                ("state", "!=", "cancel"),
                ("partner_id", "=", partner.id),
                ("ref", "=", row["comment"]),
                ("invoice_date", "=", row["invoice_date"]),
            ],
            limit=1,
        )

    def _get_or_create_partner(self, row):
        partner = self.env["res.partner"].search(
            [("vat", "=", row["cuit"])],
            limit=1,
        )
        if partner:
            return partner

        if not self.create_partners:
            raise ValidationError(
                _("No existe el cliente con CUIT %s.") % row["cuit"]
            )

        vals = {
            "name": row["partner_name"],
            "vat": row["cuit"],
            "company_type": "company",
            "customer_rank": 1,
            "country_id": self.env.ref("base.ar").id,
        }
        identification_type = self.env["l10n_latam.identification.type"].search(
            [
                ("name", "ilike", "CUIT"),
                ("country_id", "=", self.env.ref("base.ar").id),
            ],
            limit=1,
        )
        if identification_type:
            vals["l10n_latam_identification_type_id"] = identification_type.id
        return self.env["res.partner"].create(vals)

    def _get_sale_journal(self, journal_name):
        company = self.env.company
        domain = [("type", "=", "sale"), ("company_id", "=", company.id)]
        if journal_name:
            journal = self.env["account.journal"].search(
                domain + ["|", ("name", "=", journal_name), ("code", "=", journal_name)],
                limit=1,
            )
            if journal:
                return journal
            journal = self.env["account.journal"].search(
                domain + [("name", "ilike", journal_name)],
                limit=1,
            )
            if journal:
                return journal
            raise ValidationError(
                _("No se encontró un diario de ventas para '%s'.") % journal_name
            )

        journal = self.env["account.journal"].search(domain, limit=1)
        if not journal:
            raise ValidationError(_("No hay diarios de ventas configurados."))
        return journal

    def _get_payment_term(self, payment_term_name):
        if not payment_term_name:
            return False
        payment_term = self.env["account.payment.term"].search(
            [("name", "=", payment_term_name)],
            limit=1,
        )
        if payment_term:
            return payment_term
        payment_term = self.env["account.payment.term"].search(
            [("name", "ilike", payment_term_name)],
            limit=1,
        )
        if not payment_term:
            raise ValidationError(
                _("No se encontró la condición de cobro '%s'.") % payment_term_name
            )
        return payment_term

    def _get_currency(self, currency_label):
        if not currency_label or currency_label == "$":
            currency_name = "ARS"
        else:
            currency_name = currency_label
        currency = self.env["res.currency"].search(
            [("name", "=", currency_name)],
            limit=1,
        )
        return currency or self.env.company.currency_id

    def _get_sale_tax_21(self):
        return self.env["account.tax"].search(
            [
                ("type_tax_use", "=", "sale"),
                ("amount_type", "=", "percent"),
                ("amount", "=", 21.0),
                ("company_id", "=", self.env.company.id),
            ],
            limit=1,
        )

    def _get_sale_tax_exempt(self):
        return self.env["account.tax"].search(
            [
                ("type_tax_use", "=", "sale"),
                ("name", "ilike", "IVA EXENTO"),
                ("company_id", "=", self.env.company.id),
            ],
            limit=1,
        )

    def _get_default_income_account(self):
        return self.env["account.account"].search(
            [
                ("company_id", "=", self.env.company.id),
                ("deprecated", "=", False),
                ("account_type", "in", ("income", "income_other")),
            ],
            limit=1,
        )

    @staticmethod
    def _extract_code_from_header(header):
        if not header:
            return False
        match = re.search(r"\[([^\]]+)\]", str(header))
        return match.group(1).strip() if match else False

    @staticmethod
    def _cell(sheet, row, column):
        value = sheet.cell(row=row, column=column).value
        if isinstance(value, str):
            value = value.strip()
        return value

    def _to_float(self, value, allow_empty=False):
        if value in (None, ""):
            return None if allow_empty else 0.0
        if isinstance(value, (int, float)):
            return float(value)
        normalized = str(value).strip().replace(" ", "")
        if normalized.endswith("%"):
            normalized = normalized[:-1]
        if not normalized:
            return None if allow_empty else 0.0

        if "," in normalized and "." in normalized:
            if normalized.rfind(",") > normalized.rfind("."):
                normalized = normalized.replace(".", "").replace(",", ".")
            else:
                normalized = normalized.replace(",", "")
        elif "," in normalized:
            normalized = normalized.replace(",", ".")

        try:
            return float(normalized)
        except ValueError as error:
            raise ValidationError(_("Número inválido: %s") % value) from error

    def _to_date(self, value, allow_empty=False):
        if value in (None, ""):
            if allow_empty:
                return False
            raise ValidationError(_("La fecha de factura es obligatoria."))
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            # openpyxl normalmente devuelve datetime para fechas, pero por seguridad:
            return date.fromordinal(date(1899, 12, 30).toordinal() + int(value))

        raw = str(value).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        raise ValidationError(_("Fecha inválida: %s") % value)


class GcSaleInvoiceImportSheetOption(models.TransientModel):
    _name = "gc.sale.invoice.import.sheet.option"
    _description = "Opciones de hoja para importación de facturas de venta"

    wizard_id = fields.Many2one(
        "gc.sale.invoice.import.wizard",
        required=False,
        ondelete="cascade",
    )
    token = fields.Char(string="Token", required=True, index=True)
    name = fields.Char(string="Nombre de hoja", required=True)


class GcSaleInvoiceImportPreviewLine(models.TransientModel):
    _name = "gc.sale.invoice.import.preview.line"
    _description = "Preview de filas a importar de facturas de venta"
    _order = "row_number asc, id asc"

    wizard_id = fields.Many2one(
        "gc.sale.invoice.import.wizard",
        required=True,
        ondelete="cascade",
    )
    row_number = fields.Integer(string="Fila Excel", required=True)
    state = fields.Selection(
        [
            ("ready", "Lista"),
            ("duplicate", "Duplicada"),
            ("error", "Error"),
        ],
        string="Estado",
        required=True,
        default="ready",
    )
    message = fields.Char(string="Detalle")
    invoice_date = fields.Date(string="Fecha FC")
    cuit = fields.Char(string="CUIT")
    partner_name = fields.Char(string="Denominación")
    comment = fields.Char(string="Comentario")
    journal_name = fields.Char(string="Diario")
    line_amount_17 = fields.Float(string="Línea [000010]")
    line_amount_18 = fields.Float(string="Línea [000011]")
    line_amount_19 = fields.Float(string="Línea [000012]")
    vat_21 = fields.Float(string="IVA 21%")
    total = fields.Float(string="Total")
